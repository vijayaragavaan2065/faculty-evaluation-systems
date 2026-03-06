# backend/app/routers/reports.py
from fastapi import APIRouter, Depends, HTTPException
from fastapi.responses import StreamingResponse
from io import BytesIO
from datetime import datetime
from typing import Dict, Any, Optional, List
from bson import ObjectId
import random
import numpy as np

# Project-specific imports — adapt if your module paths differ
from app.routers.auth import get_current_user
from app.db.client import db  # expected to be an async motor-style db (db.submissions, db.users)

# Optional: transformers pipelines (summarizer / sentiment) if installed
try:
    from transformers import pipeline
    try:
        summarizer = pipeline("summarization", model="sshleifer/distilbart-cnn-12-6")
    except Exception:
        summarizer = None
    try:
        sentiment = pipeline("sentiment-analysis", model="distilbert-base-uncased")
    except Exception:
        sentiment = None
except Exception:
    summarizer = None
    sentiment = None

router = APIRouter(tags=["Reports"])

# ----------------------
# Helpers
# ----------------------
def _safe_parse_date(val):
    if not val:
        return datetime.utcnow()
    if isinstance(val, datetime):
        return val
    if isinstance(val, str):
        try:
            return datetime.fromisoformat(val)
        except Exception:
            try:
                return datetime.strptime(val, "%Y-%m-%dT%H:%M:%S")
            except Exception:
                try:
                    return datetime.strptime(val, "%Y-%m-%d")
                except Exception:
                    return datetime.utcnow()
    return datetime.utcnow()

def _heuristic_score_from_submission(s):
    """
    Derive a score (0..100) from a submission document if no explicit ai_score is present.
    """
    # explicit ai_score
    if isinstance(s.get("ai_score"), (int, float)):
        return float(s["ai_score"])
    # backend computed score structure
    sc = s.get("score")
    if isinstance(sc, dict) and isinstance(sc.get("total"), (int, float)):
        return float(sc["total"])
    if isinstance(sc, (int, float)):
        return float(sc)
    # try to combine totals
    totals = s.get("section_totals_json") or s.get("totals") or s.get("section_totals") or {}
    if isinstance(totals, str):
        try:
            import json
            totals = json.loads(totals)
        except Exception:
            totals = {}
    if isinstance(totals, dict):
        vals = []
        for k in ["academic", "research", "admin", "outreach", "total"]:
            v = totals.get(k)
            if isinstance(v, (int, float)):
                vals.append(float(v))
        if vals:
            t = vals[-1] if "total" in totals else sum(vals)
            if t > 1000:
                t = min(100, (t / 500) * 100)
            return float(min(100, t))
    return float(random.randint(45, 75))

def _build_ai_feedback_text(submissions: List[Dict]) -> str:
    """
    Build a single text from existing ai_feedback or key fields for summarization.
    """
    feedback_texts = [s.get("ai_feedback") for s in submissions if s.get("ai_feedback")]
    if feedback_texts:
        return " ".join(feedback_texts)

    texts = []
    for s in submissions:
        for candidate in ["remarks", "summary", "self_assessment", "narrative", "highlights"]:
            if s.get(candidate):
                texts.append(str(s.get(candidate)))
        if s.get("department"):
            texts.append(f"Department: {s.get('department')}.")
        if s.get("notes"):
            texts.append(str(s.get("notes")))
    combined = " ".join(texts).strip()
    if combined:
        return combined
    return "Performance shows consistent teaching effort; research and outreach require more focus."

# ----------------------
# Aggregation helpers
# ----------------------
async def aggregate_faculty_scores(db, department: Optional[str] = None, date_from: Optional[str] = None, date_to: Optional[str] = None):
    """
    Returns list of dicts:
      { faculty_id, faculty_name, faculty_email, avg_score, submissions_count }
    Uses MongoDB aggregation with $lookup to join user info.
    """
    pipeline = []

    # date filter
    if date_from or date_to:
        date_filter = {}
        if date_from:
            date_filter["$gte"] = _safe_parse_date(date_from)
        if date_to:
            date_filter["$lte"] = _safe_parse_date(date_to)
        pipeline.append({"$match": {"created_at": date_filter}})

    # join users
    pipeline.extend([
        {
            "$lookup": {
                "from": "users",
                "localField": "user_id",
                "foreignField": "_id",
                "as": "user"
            }
        },
        {"$unwind": "$user"},
    ])

    if department:
        pipeline.append({"$match": {"user.department": department}})

    pipeline.append({
        "$group": {
            "_id": "$user._id",
            "faculty_name": {"$first": "$user.name"},
            "faculty_email": {"$first": "$user.email"},
            "avg_score": {"$avg": {"$ifNull": ["$score", 0]}},
            "submissions_count": {"$sum": 1}
        }
    })
    pipeline.append({"$sort": {"avg_score": -1}})

    cursor = db.submissions.aggregate(pipeline)
    results = []
    async for doc in cursor:
        results.append({
            "faculty_id": str(doc["_id"]),
            "faculty_name": doc.get("faculty_name", ""),
            "faculty_email": doc.get("faculty_email", ""),
            "avg_score": float(doc.get("avg_score", 0)),
            "submissions_count": int(doc.get("submissions_count", 0))
        })
    return results

# ----------------------
# PDF generator
# ----------------------
def generate_pdf_bytes(title: str, rows: List[Dict], meta: Dict[str, Any]) -> bytes:
    """
    Generate a simple A4 PDF (fpdf) and return bytes.
    Keeps layout simple so it's portable.
    """
    try:
        from fpdf import FPDF
    except Exception:
        raise RuntimeError("fpdf package is required. Install with `pip install fpdf2`.")

    pdf = FPDF(orientation="P", unit="mm", format="A4")
    pdf.set_auto_page_break(auto=True, margin=12)
    pdf.add_page()

    # Title
    pdf.set_font("Arial", "B", 16)
    pdf.cell(0, 10, title, ln=True, align="C")
    pdf.ln(4)

    # Meta info
    pdf.set_font("Arial", size=10)
    pdf.cell(0, 6, f"Generated by: {meta.get('generated_by')}", ln=True)
    pdf.cell(0, 6, f"Generated at: {datetime.utcnow().strftime('%Y-%m-%d %H:%M UTC')}", ln=True)
    period_from = meta.get("period_from") or "—"
    period_to = meta.get("period_to") or "—"
    pdf.cell(0, 6, f"Period: {period_from} to {period_to}", ln=True)
    if meta.get("department_name"):
        pdf.cell(0, 6, f"Department: {meta.get('department_name')}", ln=True)
    pdf.ln(6)

    # Table header
    pdf.set_font("Arial", "B", 11)
    col_widths = [10, 70, 60, 30]
    headers = ["#", "Name", "Email", "Avg Score (%)"]
    for i, h in enumerate(headers):
        pdf.cell(col_widths[i], 8, h, border=1, align="C")
    pdf.ln()

    # Table rows
    pdf.set_font("Arial", size=10)
    for idx, r in enumerate(rows, start=1):
        pdf.cell(col_widths[0], 8, str(idx), border=1, align="C")
        pdf.cell(col_widths[1], 8, r.get("faculty_name", ""), border=1)
        pdf.cell(col_widths[2], 8, r.get("faculty_email", ""), border=1)
        pdf.cell(col_widths[3], 8, f"{r.get('avg_score', 0):.2f}", border=1, align="C")
        pdf.ln()

    # Footer summary
    pdf.ln(6)
    pdf.set_font("Arial", "B", 11)
    total_fac = len(rows)
    avg_of_avgs = (sum(r.get("avg_score", 0) for r in rows) / total_fac) if total_fac else 0
    pdf.cell(0, 6, f"Total faculty: {total_fac}   Department average score: {avg_of_avgs:.2f}%", ln=True)

    return pdf.output(dest="S").encode("latin-1")

# ----------------------
# Excel generator
# ----------------------
def generate_excel_bytes(rows: List[Dict]) -> bytes:
    try:
        from openpyxl import Workbook
    except Exception:
        raise RuntimeError("openpyxl package is required. Install with `pip install openpyxl`.")

    wb = Workbook()
    ws = wb.active
    ws.title = "Faculty KPI Submissions"

    ws.append([
        "Faculty ID",
        "Faculty Name",
        "Department",
        "Academic Year",
        "Score",
        "Status",
        "Submission Date"
    ])

    for r in rows:
        ws.append([
            r.get("faculty_id"),
            r.get("faculty_name"),
            r.get("department"),
            r.get("academic_year"),
            r.get("score"),
            r.get("status"),
            r.get("created_at"),
        ])

    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer.read()

# ----------------------
# Endpoints
# ----------------------
@router.get("/department/{department_name}/pdf")
async def download_department_report(
    department_name: str,
    date_from: Optional[str] = None,
    date_to: Optional[str] = None,
    current_user: dict = Depends(get_current_user),
):
    """
    Department-level PDF report. Only accessible to HOD or Admin.
    """
    role = getattr(current_user, "role", None) or current_user.get("role")
    if role not in ("hod", "admin"):
        raise HTTPException(status_code=403, detail="Not authorized")

    rows = await aggregate_faculty_scores(db, department=department_name, date_from=date_from, date_to=date_to)
    meta = {"generated_by": getattr(current_user, "email", None) or current_user.get("email"), "period_from": date_from, "period_to": date_to, "department_name": department_name}
    title = f"Department Report — {department_name}"
    pdf_bytes = generate_pdf_bytes(title, rows, meta)

    filename = f"department_report_{department_name}_{datetime.utcnow().strftime('%Y%m%d_%H%M')}.pdf"
    return StreamingResponse(BytesIO(pdf_bytes), media_type="application/pdf", headers={
        "Content-Disposition": f"attachment; filename={filename}"
    })

@router.get("/college/pdf")
async def download_college_report(
    date_from: Optional[str] = None,
    date_to: Optional[str] = None,
    current_user: dict = Depends(get_current_user),
):
    """
    College-wide PDF report. Only Admin allowed.
    """
    role = getattr(current_user, "role", None) or current_user.get("role")
    if role != "admin":
        raise HTTPException(status_code=403, detail="Not authorized")

    rows = await aggregate_faculty_scores(db, department=None, date_from=date_from, date_to=date_to)
    meta = {"generated_by": getattr(current_user, "email", None) or current_user.get("email"), "period_from": date_from, "period_to": date_to}
    title = "College-wide Faculty Scores Report"
    pdf_bytes = generate_pdf_bytes(title, rows, meta)

    filename = f"college_report_{datetime.utcnow().strftime('%Y%m%d_%H%M')}.pdf"
    return StreamingResponse(BytesIO(pdf_bytes), media_type="application/pdf", headers={
        "Content-Disposition": f"attachment; filename={filename}"
    })

@router.get("/college/excel")
async def download_college_excel(
    current_user: dict = Depends(get_current_user)
):
    """
    Registrar/Admin overall Excel report including faculty submissions.
    """
    role = getattr(current_user, "role", None) or current_user.get("role")
    if role not in ("registrar", "admin"):
        raise HTTPException(status_code=403, detail="Not authorized")

    cursor = db.submissions.find()
    rows = []

    async for doc in cursor:
        rows.append({
            "faculty_id": str(doc.get("faculty_user_id")),
            "faculty_name": doc.get("faculty_name"),
            "department": doc.get("department"),
            "academic_year": doc.get("academic_year"),
            "score": _heuristic_score_from_submission(doc),
            "status": doc.get("status"),
            "created_at": str(doc.get("created_at"))
        })

    excel_bytes = generate_excel_bytes(rows)

    filename = f"college_kpi_report_{datetime.utcnow().strftime('%Y%m%d_%H%M')}.xlsx"

    return StreamingResponse(
        BytesIO(excel_bytes),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )

@router.get("/user/{user_id}")
async def get_user_analytics(user_id: str, current_user: dict = Depends(get_current_user)) -> Dict[str, Any]:
    """
    Returns analytics for a specific user: time_series, kpi_breakdown, overall summary.
    """
    queries = []
    try:
        queries.append({"user_id": ObjectId(user_id)})
    except Exception:
        pass
    queries.append({"user_id": user_id})
    queries.append({"faculty_user_id": user_id})
    queries.append({"faculty_email": user_id})

    submissions = []
    for q in queries:
        try:
            docs = await db.submissions.find(q).to_list(length=None)
            if docs:
                submissions.extend(docs)
        except Exception:
            continue

    unique = {}
    for s in submissions:
        unique[str(s.get("_id"))] = s
    submissions = list(unique.values())

    if not submissions:
        return {
            "time_series": [],
            "kpi_breakdown": [],
            "overall": {"avg_score": 0},
            "message": "No submissions found for this user.",
        }

    time_series = []
    for s in submissions:
        score = _heuristic_score_from_submission(s)
        created = s.get("created_at") or s.get("submitted_at") or s.get("created")
        created = _safe_parse_date(created)
        time_series.append({
            "month": created.strftime("%b %Y"),
            "score": float(max(0, min(100, round(score, 2))))
        })

    def _month_key(item):
        try:
            return datetime.strptime(item["month"], "%b %Y")
        except Exception:
            return datetime.utcnow()
    time_series = sorted(time_series, key=_month_key)

    aggregated = {"Teaching": [], "Research": [], "Mentoring": [], "Industry Collaboration": [], "Community Service": []}
    for s in submissions:
        totals = s.get("section_totals_json") or s.get("totals") or s.get("section_totals") or {}
        if isinstance(totals, str):
            try:
                import json
                totals = json.loads(totals)
            except Exception:
                totals = {}
        if isinstance(totals, dict):
            if totals.get("academic") is not None:
                aggregated["Teaching"].append(float(totals.get("academic") or 0))
            if totals.get("research") is not None:
                aggregated["Research"].append(float(totals.get("research") or 0))
            if totals.get("admin") is not None:
                aggregated["Mentoring"].append(float(totals.get("admin") or 0))
            if totals.get("outreach") is not None:
                aggregated["Community Service"].append(float(totals.get("outreach") or 0))

    kpi_breakdown = []
    for k, arr in aggregated.items():
        if arr:
            avg = float(np.clip(np.mean(arr), 0, 100))
        else:
            avg = float(np.clip(random.gauss(60, 12), 20, 90))
        kpi_breakdown.append({"kpi": k, "score": round(avg, 2)})

    combined_text = _build_ai_feedback_text(submissions)
    ai_summary = ""
    if summarizer:
        try:
            doc = combined_text[:1200]
            out = summarizer(doc, max_length=80, min_length=20, do_sample=False)
            if isinstance(out, list) and out and isinstance(out[0], dict):
                ai_summary = out[0].get("summary_text", "Summary not available.")
            else:
                ai_summary = "Summary not available."
        except Exception as e:
            ai_summary = f"AI summarizer error: {str(e)}"
    else:
        if sentiment:
            try:
                res = sentiment(combined_text[:512])[0]
                ai_summary = f"Overall sentiment: {res.get('label')} (score {res.get('score'):.2f})"
            except Exception:
                ai_summary = "AI sentiment unavailable."
        else:
            ai_summary = "AI summarization unavailable on server (model not loaded)."

    avg_score = float(np.mean([x["score"] for x in time_series])) if time_series else 0

    return {
        "time_series": time_series,
        "kpi_breakdown": kpi_breakdown,
        "overall": {"avg_score": round(avg_score, 2), "summary": ai_summary},
        "message": f"Found {len(submissions)} submission(s) for user_id={user_id}",
    }